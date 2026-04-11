"""
Excel Export — Generates formatted .xlsx report from detection history
"""
import os
from datetime import datetime
from typing import Any

try:
    from openpyxl import Workbook
    from openpyxl.styles import (
        PatternFill, Font, Alignment, Border, Side, GradientFill
    )
    from openpyxl.chart import BarChart, Reference
    from openpyxl.chart.series import DataPoint
    from openpyxl.utils import get_column_letter
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False
    print("⚠️  openpyxl not installed — Excel export disabled")


# ─── Colour Palette ───────────────────────────────────────────────────────────
PALETTE = {
    "header_bg":    "1A1A2E",
    "header_fg":    "FF6B35",
    "row_even":     "F8F9FA",
    "row_odd":      "FFFFFF",
    "high_bg":      "FF4444",
    "high_fg":      "FFFFFF",
    "medium_bg":    "FF9800",
    "medium_fg":    "FFFFFF",
    "low_bg":       "4CAF50",
    "low_fg":       "FFFFFF",
    "accent":       "FF6B35",
    "border":       "CCCCCC",
}

COLS = [
    ("Date",       "date",      14),
    ("Time",       "time",      10),
    ("Label",      "label",     14),
    ("Confidence", "conf",      12),
    ("Severity",   "severity",  12),
    ("Source",     "source",    16),
    ("Timestamp",  "timestamp", 22),
]


def export_detections_to_excel(records: list[dict]) -> str:
    """
    Creates an Excel workbook with:
      - Sheet 1: Full detection log with formatting
      - Sheet 2: Summary statistics
      - Sheet 3: Bar chart of label frequencies
    Returns the file path.
    """
    os.makedirs("exports", exist_ok=True)
    filepath = f"exports/detections_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

    if not OPENPYXL_AVAILABLE:
        # Fallback: plain CSV
        csv_path = filepath.replace(".xlsx", ".csv")
        with open(csv_path, "w") as f:
            f.write(",".join(h for h, _, _ in COLS) + "\n")
            for r in records:
                row = [str(r.get(k, "")) for _, k, _ in COLS]
                f.write(",".join(row) + "\n")
        return csv_path

    wb = Workbook()

    # ── Sheet 1 : Detection Log ────────────────────────────────────────────
    ws = wb.active
    ws.title = "Detection Log"

    _write_title(ws, "🔥 AI Surveillance — Detection History")
    _write_meta(ws, len(records))
    _write_headers(ws, start_row=4)
    _write_data_rows(ws, records, start_row=5)
    _set_col_widths(ws)
    _freeze_panes(ws)
    _add_autofilter(ws, len(records))

    # ── Sheet 2 : Summary ─────────────────────────────────────────────────
    ws2 = wb.create_sheet("Summary")
    _write_summary_sheet(ws2, records)

    # ── Sheet 3 : Chart ───────────────────────────────────────────────────
    ws3 = wb.create_sheet("Charts")
    _write_chart_sheet(ws3, records, wb)

    wb.save(filepath)
    return filepath


# ─── Helpers ──────────────────────────────────────────────────────────────────
def _style_cell(cell, bg=None, fg="000000", bold=False,
                align="left", font_size=11):
    if bg:
        cell.fill = PatternFill("solid", fgColor=bg)
    cell.font = Font(bold=bold, color=fg, size=font_size,
                     name="Calibri")
    cell.alignment = Alignment(horizontal=align, vertical="center",
                               wrap_text=False)


def _border():
    side = Side(style="thin", color=PALETTE["border"])
    return Border(left=side, right=side, top=side, bottom=side)


def _write_title(ws, title: str):
    ws.merge_cells("A1:G1")
    cell = ws["A1"]
    cell.value = title
    cell.fill  = PatternFill("solid", fgColor=PALETTE["header_bg"])
    cell.font  = Font(bold=True, size=16, color=PALETTE["header_fg"],
                      name="Calibri")
    cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 36


def _write_meta(ws, total: int):
    ws.merge_cells("A2:G2")
    cell = ws["A2"]
    cell.value = (
        f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')} "
        f"  |  Total Records: {total}"
    )
    cell.font      = Font(italic=True, size=10, color="666666", name="Calibri")
    cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 20
    ws.row_dimensions[3].height = 6  # spacer


def _write_headers(ws, start_row: int):
    for col_idx, (header, _, _) in enumerate(COLS, start=1):
        cell = ws.cell(row=start_row, column=col_idx, value=header)
        _style_cell(cell, bg=PALETTE["header_bg"], fg=PALETTE["header_fg"],
                    bold=True, align="center", font_size=11)
        cell.border = _border()
    ws.row_dimensions[start_row].height = 28


def _write_data_rows(ws, records: list[dict], start_row: int):
    for row_idx, record in enumerate(records):
        actual_row = start_row + row_idx
        bg = PALETTE["row_even"] if row_idx % 2 == 0 else PALETTE["row_odd"]
        severity = record.get("severity", "MEDIUM")

        for col_idx, (_, key, _) in enumerate(COLS, start=1):
            value = record.get(key, "")
            if key == "conf" and isinstance(value, float):
                value = f"{value:.1%}"

            cell = ws.cell(row=actual_row, column=col_idx, value=value)
            cell.border = _border()

            if key == "severity":
                s_bg = {
                    "HIGH":   PALETTE["high_bg"],
                    "MEDIUM": PALETTE["medium_bg"],
                    "LOW":    PALETTE["low_bg"],
                }.get(severity, PALETTE["medium_bg"])
                _style_cell(cell, bg=s_bg, fg="FFFFFF", align="center")
            else:
                _style_cell(cell, bg=bg, align="left")

        ws.row_dimensions[actual_row].height = 20


def _set_col_widths(ws):
    for col_idx, (_, _, width) in enumerate(COLS, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width


def _freeze_panes(ws):
    ws.freeze_panes = "A5"


def _add_autofilter(ws, total: int):
    ws.auto_filter.ref = f"A4:G{4 + total}"


def _write_summary_sheet(ws, records: list[dict]):
    ws["A1"] = "📊 Detection Summary"
    ws["A1"].font = Font(bold=True, size=14, color=PALETTE["accent"], name="Calibri")

    # Label counts
    label_counts: dict[str, int] = {}
    severity_counts: dict[str, int] = {}
    date_counts: dict[str, int] = {}

    for r in records:
        lbl  = r.get("label", "unknown")
        sev  = r.get("severity", "MEDIUM")
        date = r.get("date", "unknown")
        label_counts[lbl]    = label_counts.get(lbl, 0) + 1
        severity_counts[sev] = severity_counts.get(sev, 0) + 1
        date_counts[date]    = date_counts.get(date, 0) + 1

    row = 3
    ws.cell(row=row, column=1, value="Label").font = Font(bold=True, name="Calibri")
    ws.cell(row=row, column=2, value="Count").font = Font(bold=True, name="Calibri")
    for lbl, cnt in sorted(label_counts.items(), key=lambda x: -x[1]):
        row += 1
        ws.cell(row=row, column=1, value=lbl)
        ws.cell(row=row, column=2, value=cnt)

    row += 2
    ws.cell(row=row, column=1, value="Severity").font = Font(bold=True, name="Calibri")
    ws.cell(row=row, column=2, value="Count").font   = Font(bold=True, name="Calibri")
    for sev, cnt in severity_counts.items():
        row += 1
        ws.cell(row=row, column=1, value=sev)
        ws.cell(row=row, column=2, value=cnt)

    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["B"].width = 12


def _write_chart_sheet(ws, records: list[dict], wb):
    ws["A1"] = "Label"
    ws["B1"] = "Detections"

    label_counts: dict[str, int] = {}
    for r in records:
        lbl = r.get("label", "unknown")
        label_counts[lbl] = label_counts.get(lbl, 0) + 1

    for i, (lbl, cnt) in enumerate(
        sorted(label_counts.items(), key=lambda x: -x[1])[:15], start=2
    ):
        ws.cell(row=i, column=1, value=lbl)
        ws.cell(row=i, column=2, value=cnt)

    n = len(label_counts) + 1
    chart = BarChart()
    chart.type  = "col"
    chart.title = "Detections by Label"
    chart.style = 10
    chart.y_axis.title = "Count"
    chart.x_axis.title = "Label"

    data = Reference(ws, min_col=2, min_row=1, max_row=n)
    cats = Reference(ws, min_col=1, min_row=2, max_row=n)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    chart.width  = 24
    chart.height = 14

    ws.add_chart(chart, "D2")
    ws.column_dimensions["A"].width = 18
    ws.column_dimensions["B"].width = 14
